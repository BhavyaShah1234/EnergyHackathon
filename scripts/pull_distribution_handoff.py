#!/usr/bin/env python3
"""Pull broad-coverage handoff data for the chart items used in siting.

This script is intentionally about source coverage, not model training.

It exports the chart branches we can reach from public sources today:
  - land/legal-ish context: BLM SMA + Texas GLO lease datasets
  - water/flood: NHD waterbody + FEMA floodplain
  - pipeline access: EIA pipeline geometry
  - grid prices: CAISO OASIS + ERCOT historical DAM/RTM hub/load-zone prices
  - fuel prices: Henry Hub daily history
  - grid conditions: EIA-930 BA history

It also writes a manifest that calls out the remaining blocked gaps:
  - PHMSA incidents / annual reports
  - EIA-176 / EIA-757
  - Waha history
"""
from __future__ import annotations

import csv
import io
import json
import os
import sys
import time
import zipfile
from datetime import date, datetime, time as dtime, timedelta, timezone
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from zoneinfo import ZoneInfo


ROOT = Path(__file__).resolve().parent.parent
OUT_ROOT = ROOT / "data" / "training"
NOW = datetime.now(timezone.utc)
TS_TAG = NOW.strftime("%Y%m%dT%H%M%SZ")
OUT_DIR = OUT_ROOT / f"distribution_handoff_{TS_TAG}"
OUT_DIR.mkdir(parents=True, exist_ok=True)

NOAA_UA = os.environ.get("NOAA_USER_AGENT", "collide-energy-pipeline sdonthi4@asu.edu")
EIA_KEY = ""
ERCOT_TZ = ZoneInfo("America/Chicago")

CAISO_NODES = ["PALOVRDE_ASR-APND", "TH_SP15_GEN-APND", "TH_NP15_GEN-APND"]
ARC_PAGE = 2000

GLO_DATASETS = {
    "glo_upland_leases": {
        "url": "https://services1.arcgis.com/YWG34dhJxrbxQWdF/arcgis/rest/services/Upland_Leases/FeatureServer/0/query",
        "fields": [
            "OBJECTID_1", "PROJECT_NUMBER", "PROJECT_NAME", "LEASE_NUMBER", "LEASE_STATUS",
            "ACTIVITY", "TOTAL_CONSIDERATION", "PRIMARY_LESSEE", "ALL_LESSEE",
            "PROJECT_LATITUDE", "PROJECT_LONGITUDE", "PurposeClass", "GLOID",
        ],
    },
    "glo_oilgas_active": {
        "url": "https://services1.arcgis.com/YWG34dhJxrbxQWdF/arcgis/rest/services/Oil_and_Gas_Leases_Active/FeatureServer/0/query",
        "fields": [
            "OBJECTID", "LEASE_NUMBER", "LEASE_STATUS", "LEASE_STATUS_DATE", "EFFECTIVE_DATE",
            "PRIMARY_TERM_END_DATE", "ORIGINAL_GROSS_ACRES", "CURRENT_NET_ACRES",
            "LEASE_TYPE", "ORIGINAL_LESSEE", "LESSOR", "COUNTY", "LEASE_ROYALTY_GAS",
            "LEASE_ROYALTY_OIL", "LAND_TYPE", "FIRST_WELL_CLASS", "LEASE_UPDATE",
        ],
    },
    "glo_oilgas_inactive": {
        "url": "https://services1.arcgis.com/YWG34dhJxrbxQWdF/arcgis/rest/services/Oil_and_Gas_Leases_In_Active/FeatureServer/0/query",
        "fields": [
            "OBJECTID", "LEASE_NUMBER", "LEASE_STATUS", "LEASE_STATUS_DATE", "EFFECTIVE_DATE",
            "PRIMARY_TERM_END_DATE", "ORIGINAL_GROSS_ACRES", "CURRENT_NET_ACRES",
            "LEASE_TYPE", "ORIGINAL_LESSEE", "LESSOR", "COUNTY", "LEASE_ROYALTY_GAS",
            "LEASE_ROYALTY_OIL", "LAND_TYPE", "FIRST_WELL_CLASS", "LEASE_UPDATE",
        ],
    },
}


def _load_env() -> None:
    env_path = ROOT / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, value = line.split("=", 1)
                os.environ.setdefault(key.strip(), value.strip())


def _get(url: str, headers: dict | None = None, timeout: int = 90) -> bytes:
    hdrs = {"User-Agent": NOAA_UA, "Accept": "*/*"}
    if headers:
        hdrs.update(headers)
    req = Request(url, headers=hdrs)
    last_err = None
    for attempt in range(4):
        try:
            with urlopen(req, timeout=timeout) as resp:
                return resp.read()
        except Exception as exc:  # pragma: no cover - network retry wrapper
            last_err = exc
            if attempt == 3:
                raise
            time.sleep(2 ** attempt)
    raise last_err  # pragma: no cover


def _save_csv(dataset: str, rows: list[dict]) -> Path:
    path = OUT_DIR / f"{dataset}.csv"
    if not rows:
        path.write_text("")
        return path
    cols = sorted(set().union(*(r.keys() for r in rows)))
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=cols)
        writer.writeheader()
        writer.writerows(rows)
    return path


def _save_parquet_if_possible(dataset: str, rows: list[dict]) -> Path | None:
    if not rows:
        return None
    try:
        import pandas as pd

        path = OUT_DIR / f"{dataset}.parquet"
        pd.DataFrame(rows).to_parquet(path, index=False, engine="pyarrow")
        return path
    except Exception:
        return None


def _json_from_url(base: str, params: dict[str, str], timeout: int = 90) -> dict:
    body = _get(f"{base}?{urlencode(params)}", timeout=timeout)
    return json.loads(body)


def _paginate_arcgis_geojson(base: str, params: dict[str, str], page_size: int = ARC_PAGE) -> list[dict]:
    rows: list[dict] = []
    offset = 0
    while True:
        page = dict(params)
        page["resultOffset"] = str(offset)
        page["resultRecordCount"] = str(page_size)
        payload = _json_from_url(base, page)
        features = payload.get("features", [])
        rows.extend(features)
        if len(features) < page_size:
            break
        offset += page_size
    return rows


def pull_blm_sma() -> list[dict]:
    params = {
        "where": "ADMIN_ST IN ('AZ','NM','TX')",
        "outFields": "OBJECTID,SMA_ID,ADMIN_DEPT_CODE,ADMIN_AGENCY_CODE,ADMIN_UNIT_NAME,ADMIN_UNIT_TYPE,ADMIN_ST,SHAPE_Area",
        "f": "json",
        "returnGeometry": "false",
    }
    features = _paginate_arcgis_geojson(
        "https://gis.blm.gov/arcgis/rest/services/lands/BLM_Natl_SMA_Cached_without_PriUnk/MapServer/1/query",
        params,
    )
    rows = []
    for feat in features:
        props = feat.get("attributes", {})
        rows.append({
            "_source": "blm_sma",
            "dataset": "blm_sma",
            "object_id": props.get("OBJECTID"),
            "sma_id": props.get("SMA_ID"),
            "admin_department": props.get("ADMIN_DEPT_CODE"),
            "admin_agency": props.get("ADMIN_AGENCY_CODE"),
            "admin_state": props.get("ADMIN_ST"),
            "admin_unit_name": props.get("ADMIN_UNIT_NAME"),
            "admin_unit_type": props.get("ADMIN_UNIT_TYPE"),
            "shape_area_sq_deg": props.get("SHAPE_Area"),
        })
    return rows


def pull_glo_dataset(dataset: str, config: dict) -> list[dict]:
    params = {
        "where": "1=1",
        "outFields": ",".join(config["fields"]),
        "f": "geojson",
        "returnGeometry": "true",
        "outSR": "4326",
    }
    features = _paginate_arcgis_geojson(config["url"], params)
    rows = []
    for feat in features:
        props = feat.get("properties", {})
        row = {"_source": dataset, "dataset": dataset, "geometry_geojson": json.dumps(feat.get("geometry"))}
        for key, value in props.items():
            row[key.lower()] = value
        rows.append(row)
    return rows


def pull_nhd_waterbody() -> list[dict]:
    boxes = [
        "-114.82,31.33,-109.04,37.00",
        "-109.05,31.33,-103.00,37.00",
        "-106.65,25.84,-93.51,36.50",
    ]
    rows = []
    for bbox in boxes:
        params = {
            "where": "1=1",
            "geometry": bbox,
            "geometryType": "esriGeometryEnvelope",
            "spatialRel": "esriSpatialRelIntersects",
            "outFields": "OBJECTID,GNIS_NAME,FTYPE,FCODE,AREASQKM,REACHCODE",
            "f": "geojson",
            "returnGeometry": "true",
            "inSR": "4326",
            "outSR": "4326",
        }
        for feat in _paginate_arcgis_geojson(
            "https://hydro.nationalmap.gov/arcgis/rest/services/nhd/MapServer/10/query",
            params,
        ):
            props = feat.get("properties", {})
            rows.append({
                "_source": "nhd_waterbody",
                "dataset": "nhd_waterbody",
                "object_id": props.get("OBJECTID"),
                "gnis_name": props.get("GNIS_NAME"),
                "feature_type": props.get("FTYPE"),
                "feature_code": props.get("FCODE"),
                "area_sq_km": props.get("AREASQKM"),
                "reach_code": props.get("REACHCODE"),
                "geometry_geojson": json.dumps(feat.get("geometry")),
            })
    return rows


def pull_fema_floodplain() -> list[dict]:
    boxes = [
        "-114.82,31.33,-109.04,37.00",
        "-109.05,31.33,-103.00,37.00",
        "-106.65,25.84,-93.51,36.50",
    ]
    rows = []
    for bbox in boxes:
        params = {
            "where": "1=1",
            "geometry": bbox,
            "geometryType": "esriGeometryEnvelope",
            "spatialRel": "esriSpatialRelIntersects",
            "outFields": "OBJECTID,FLD_AR_ID,FLD_ZONE,ZONE_SUBTY,SFHA_TF,STATIC_BFE,DEPTH",
            "f": "geojson",
            "returnGeometry": "true",
            "inSR": "4326",
            "outSR": "4326",
        }
        for feat in _paginate_arcgis_geojson(
            "https://hazards.fema.gov/arcgis/rest/services/public/NFHL/MapServer/28/query",
            params,
            page_size=250,
        ):
            props = feat.get("properties", {})
            rows.append({
                "_source": "fema_floodplain",
                "dataset": "fema_floodplain",
                "object_id": props.get("OBJECTID"),
                "flood_area_id": props.get("FLD_AR_ID"),
                "flood_zone": props.get("FLD_ZONE"),
                "zone_subtype": props.get("ZONE_SUBTY"),
                "sfha_flag": props.get("SFHA_TF"),
                "static_bfe_ft": props.get("STATIC_BFE"),
                "depth_ft": props.get("DEPTH"),
                "geometry_geojson": json.dumps(feat.get("geometry")),
            })
    return rows


def pull_pipelines_infra() -> list[dict]:
    offset = 0
    rows = []
    while True:
        params = {
            "where": "1=1",
            "outFields": "FID,TYPEPIPE,Operator,Status",
            "returnGeometry": "true",
            "outSR": "4326",
            "geometry": json.dumps({"xmin": -125.0, "ymin": 25.0, "xmax": -93.0, "ymax": 42.0}),
            "geometryType": "esriGeometryEnvelope",
            "inSR": "4326",
            "spatialRel": "esriSpatialRelIntersects",
            "resultOffset": str(offset),
            "resultRecordCount": str(ARC_PAGE),
            "f": "json",
        }
        payload = _json_from_url(
            "https://services2.arcgis.com/FiaPA4ga0iQKduv3/arcgis/rest/services/Natural_Gas_Interstate_and_Intrastate_Pipelines_1/FeatureServer/0/query",
            params,
        )
        features = payload.get("features", [])
        for feat in features:
            attrs = feat.get("attributes", {})
            rows.append({
                "_source": "pipelines_infra",
                "dataset": "pipelines_infra",
                "pipeline_id": attrs.get("FID"),
                "pipe_type": attrs.get("TYPEPIPE"),
                "operator": attrs.get("Operator"),
                "status": attrs.get("Status"),
                "geometry_json": json.dumps(feat.get("geometry")),
            })
        if not payload.get("exceededTransferLimit"):
            break
        offset += ARC_PAGE
    return rows


def pull_eia930(days: int = 365) -> list[dict]:
    if not EIA_KEY:
        return []
    start = (NOW - timedelta(days=days)).strftime("%Y-%m-%d")
    rows = []
    for ba in ["AZPS", "CISO", "ERCO"]:
        offset = 0
        while True:
            url = (
                "https://api.eia.gov/v2/electricity/rto/region-data/data/"
                f"?api_key={EIA_KEY}&frequency=hourly&data[0]=value"
                f"&facets[respondent][]={ba}"
                "&facets[type][]=D&facets[type][]=DF&facets[type][]=NG&facets[type][]=TI"
                f"&start={start}"
                "&sort[0][column]=period&sort[0][direction]=desc"
                f"&length=5000&offset={offset}"
            )
            data = json.loads(_get(url))
            records = data.get("response", {}).get("data", [])
            for record in records:
                value = record.get("value")
                rows.append({
                    "_source": f"eia930_{ba.lower()}",
                    "dataset": "eia930",
                    "period_utc": record.get("period"),
                    "respondent": record.get("respondent"),
                    "respondent_name": record.get("respondent-name"),
                    "type": record.get("type"),
                    "type_name": record.get("type-name"),
                    "value_mw": float(value) if value is not None else None,
                })
            if len(records) < 5000:
                break
            offset += 5000
            time.sleep(0.2)
    return rows


def pull_eia_ng_henry(days: int = 365) -> list[dict]:
    if not EIA_KEY:
        return []
    start = (NOW - timedelta(days=days)).strftime("%Y-%m-%d")
    url = (
        "https://api.eia.gov/v2/natural-gas/pri/fut/data/"
        f"?api_key={EIA_KEY}&frequency=daily&data[0]=value"
        "&facets[series][]=RNGWHHD"
        f"&start={start}"
        "&sort[0][column]=period&sort[0][direction]=desc"
        "&length=5000"
    )
    data = json.loads(_get(url))
    rows = []
    for record in data.get("response", {}).get("data", []):
        value = record.get("value")
        rows.append({
            "_source": "eia_ng_henry_hub",
            "dataset": "eia_ng_henry_hub",
            "period_utc": record.get("period"),
            "series": record.get("series"),
            "series_description": record.get("series-description"),
            "price_usd_per_mmbtu": float(value) if value is not None else None,
        })
    return rows


def pull_caiso_lmp(days: int = 30) -> list[dict]:
    rows = []
    end = NOW.replace(second=0, microsecond=0)
    end = end.replace(minute=(end.minute // 5) * 5)
    start = end - timedelta(days=days)
    for node in CAISO_NODES:
        params = (
            "queryname=PRC_INTVL_LMP&market_run_id=RTM&version=1&resultformat=6"
            f"&node={node}"
            f"&startdatetime={start.strftime('%Y%m%dT%H:%M-0000')}"
            f"&enddatetime={end.strftime('%Y%m%dT%H:%M-0000')}"
        )
        body = _get(f"https://oasis.caiso.com/oasisapi/SingleZip?{params}", timeout=180)
        zf = zipfile.ZipFile(io.BytesIO(body))
        for name in zf.namelist():
            if not name.endswith(".csv"):
                continue
            with zf.open(name) as fh:
                reader = csv.DictReader(io.TextIOWrapper(fh))
                for row in reader:
                    value = row.get("VALUE") or row.get("MW")
                    rows.append({
                        "_source": "caiso_lmp",
                        "dataset": "caiso_lmp",
                        "interval_start_utc": row.get("INTERVALSTARTTIME_GMT"),
                        "interval_end_utc": row.get("INTERVALENDTIME_GMT"),
                        "node": row.get("NODE"),
                        "lmp_component": row.get("LMP_TYPE"),
                        "price_usd_per_mwh": float(value) if value else None,
                    })
        time.sleep(1.0)
    return rows


def _iter_ercot_docs(report_type_id: int) -> list[dict]:
    payload = json.loads(_get(f"https://www.ercot.com/misapp/servlets/IceDocListJsonWS?reportTypeId={report_type_id}"))
    docs = payload["ListDocsByRptTypeRes"]["DocumentList"]
    items = docs if isinstance(docs, list) else [docs]
    return [item["Document"] for item in items]


def _read_ercot_xlsx_rows(blob: bytes):
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    iterator = worksheet.iter_rows(values_only=True)
    header = next(iterator)
    for values in iterator:
        yield dict(zip(header, values))


def pull_ercot_dam() -> list[dict]:
    rows = []
    for doc in _iter_ercot_docs(13060):
        zip_blob = _get(f"https://www.ercot.com/misdownload/servlets/mirDownload?doclookupId={doc['DocID']}", timeout=180)
        zf = zipfile.ZipFile(io.BytesIO(zip_blob))
        sheet_blob = zf.read(zf.namelist()[0])
        for raw in _read_ercot_xlsx_rows(sheet_blob):
            delivery_date = datetime.strptime(raw["Delivery Date"], "%m/%d/%Y").date()
            hour_ending = int(str(raw["Hour Ending"]).split(":")[0])
            start_local = datetime.combine(delivery_date, dtime.min, tzinfo=ERCOT_TZ) + timedelta(hours=hour_ending - 1)
            rows.append({
                "_source": "ercot_dam_hub_prices",
                "dataset": "ercot_dam_hub_prices",
                "delivery_date_local": delivery_date.isoformat(),
                "hour_ending": raw["Hour Ending"],
                "repeated_hour_flag": raw["Repeated Hour Flag"],
                "settlement_point": raw["Settlement Point"],
                "price_usd_per_mwh": raw["Settlement Point Price"],
                "interval_start_utc": start_local.astimezone(timezone.utc).isoformat(),
                "report_friendly_name": doc["FriendlyName"],
            })
    return rows


def pull_ercot_rtm() -> list[dict]:
    rows = []
    for doc in _iter_ercot_docs(13061):
        zip_blob = _get(f"https://www.ercot.com/misdownload/servlets/mirDownload?doclookupId={doc['DocID']}", timeout=180)
        zf = zipfile.ZipFile(io.BytesIO(zip_blob))
        sheet_blob = zf.read(zf.namelist()[0])
        for raw in _read_ercot_xlsx_rows(sheet_blob):
            delivery_date = datetime.strptime(raw["Delivery Date"], "%m/%d/%Y").date()
            delivery_hour = int(raw["Delivery Hour"])
            interval = int(raw["Delivery Interval"])
            start_local = datetime.combine(delivery_date, dtime.min, tzinfo=ERCOT_TZ)
            start_local += timedelta(hours=delivery_hour - 1, minutes=(interval - 1) * 15)
            rows.append({
                "_source": "ercot_rtm_hub_prices",
                "dataset": "ercot_rtm_hub_prices",
                "delivery_date_local": delivery_date.isoformat(),
                "delivery_hour": delivery_hour,
                "delivery_interval": interval,
                "repeated_hour_flag": raw["Repeated Hour Flag"],
                "settlement_point_name": raw["Settlement Point Name"],
                "settlement_point_type": raw["Settlement Point Type"],
                "price_usd_per_mwh": raw["Settlement Point Price"],
                "interval_start_utc": start_local.astimezone(timezone.utc).isoformat(),
                "report_friendly_name": doc["FriendlyName"],
            })
    return rows


def main() -> int:
    _load_env()
    global EIA_KEY
    EIA_KEY = os.environ.get("EIA_API_KEY", "")

    print(f"Distribution handoff pull -> {OUT_DIR}")
    print(f"UTC timestamp: {NOW.isoformat()}")

    pullers = [
        ("blm_sma", pull_blm_sma),
        ("glo_upland_leases", lambda: pull_glo_dataset("glo_upland_leases", GLO_DATASETS["glo_upland_leases"])),
        ("glo_oilgas_active", lambda: pull_glo_dataset("glo_oilgas_active", GLO_DATASETS["glo_oilgas_active"])),
        ("glo_oilgas_inactive", lambda: pull_glo_dataset("glo_oilgas_inactive", GLO_DATASETS["glo_oilgas_inactive"])),
        ("nhd_waterbody", pull_nhd_waterbody),
        ("fema_floodplain", pull_fema_floodplain),
        ("pipelines_infra", pull_pipelines_infra),
        ("eia930", pull_eia930),
        ("eia_ng_henry_hub", pull_eia_ng_henry),
        ("caiso_lmp", pull_caiso_lmp),
        ("ercot_dam_hub_prices", pull_ercot_dam),
        ("ercot_rtm_hub_prices", pull_ercot_rtm),
    ]

    if "--only" in sys.argv:
        idx = sys.argv.index("--only")
        wanted = set(sys.argv[idx + 1].split(",")) if idx + 1 < len(sys.argv) else set()
        pullers = [item for item in pullers if item[0] in wanted]

    manifest = {
        "generated_at_utc": NOW.isoformat(),
        "output_dir": str(OUT_DIR),
        "datasets": {},
        "blocked_or_missing": {
            "phmsa_incidents": "Official PHMSA bulk download returned HTTP 403 during automation.",
            "phmsa_annual_reports": "Official PHMSA bulk download returned HTTP 403 during automation.",
            "eia_176_757": "Official EIA query-system direct bulk path not wired yet.",
            "waha_history": "Current EIA API series config does not return Waha spot history.",
        },
    }

    for dataset, fn in pullers:
        print(f"Pulling {dataset}...", flush=True)
        try:
            rows = fn()
            csv_path = _save_csv(dataset, rows)
            parquet_path = _save_parquet_if_possible(dataset, rows)
            manifest["datasets"][dataset] = {
                "row_count": len(rows),
                "csv": csv_path.name,
                "parquet": parquet_path.name if parquet_path else None,
                "error": None,
            }
            print(f"  -> {len(rows)} rows")
        except Exception as exc:
            manifest["datasets"][dataset] = {
                "row_count": 0,
                "csv": None,
                "parquet": None,
                "error": f"{type(exc).__name__}: {exc}",
            }
            print(f"  -> FAILED: {type(exc).__name__}: {exc}")

    manifest_path = OUT_DIR / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2))
    print(f"Manifest -> {manifest_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
